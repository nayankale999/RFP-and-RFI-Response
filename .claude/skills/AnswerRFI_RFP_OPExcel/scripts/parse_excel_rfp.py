#!/usr/bin/env python3
"""
RFP/RFI Excel Questionnaire Parser and Answer Writer

Parses Excel-based RFP/RFI questionnaires, extracts questions as structured JSON,
and writes generated answers back into the vendor response column.

Usage:
    # List sheets:
    python3 parse_excel_rfp.py --input rfp.xlsx --parse-only

    # Extract questions from specific sheets:
    python3 parse_excel_rfp.py --input rfp.xlsx --parse-only --sheets "D. Functional Requirements,E. Non-Functional"

    # Write answers back:
    python3 parse_excel_rfp.py --input rfp.xlsx --write-answers answers.json [--output-file output.xlsx]
"""

import argparse
import json
import logging
import os
import re
import sys
from datetime import datetime
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, column_index_from_string

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Header keyword patterns for column detection (case-insensitive)
RESPONSE_KEYWORDS = [
    "vendor response", "response", "vendor answer", "answer",
    "supplier response", "your response", "bidder response",
]
QUESTION_KEYWORDS = [
    "question", "requirement", "description", "criteria",
    "request", "query", "detail", "specification",
]
ID_KEYWORDS = [
    "id", "ref", "reference", "#", "no.", "number", "item", "sr",
]
SCORE_KEYWORDS = [
    "vendor score", "score", "rating", "compliance", "vendor rating",
]

# Question type classification patterns
BINARY_PATTERNS = [
    r"(?i)^The system\b",
    r"(?i)^The solution\b",
    r"(?i)^The platform\b",
    r"(?i)^The product\b",
    r"(?i)^Does (your|the) (system|solution|platform|product)\b",
    r"(?i)^Can (your|the) (system|solution|platform|product|users?)\b",
    r"(?i)^Is (your|the) (system|solution|platform|product)\b",
    r"(?i)^Are (you|users|administrators)\b.*able to\b",
    r"(?i)\b(Y/N|Yes/No)\b",
    r"(?i)^(Your|The) (system|solution|platform|product) (supports?|allows?|enables?|provides?|includes?|offers?|has)\b",
    r"(?i)^It is possible\b",
    r"(?i)^There is (a |an )?\b",
]
NARRATIVE_PATTERNS = [
    r"(?i)^(Describe|Explain|Provide|Detail|Outline|Elaborate|Specify|List|Summarize)",
    r"(?i)^(How does|How do|How can|How will|How would)",
    r"(?i)^(What is|What are|What does|What will)",
    r"(?i)^(Please (describe|explain|provide|detail|list|outline|specify))",
    r"(?i)^(Give|State|Indicate|Identify|Define|Clarify)",
]
COMPANY_INFO_PATTERNS = [
    r"(?i)(company name|organisation name|organization name)",
    r"(?i)(headquarters|head office|registered address)",
    r"(?i)(number of employees|headcount|staff count|employee count)",
    r"(?i)(annual revenue|turnover|financial)",
    r"(?i)(year (founded|established|incorporated))",
    r"(?i)(ownership (structure|type))",
    r"(?i)(CEO|CTO|managing director|board of directors)",
    r"(?i)(parent company|subsidiary)",
    r"(?i)(stock|ticker|publicly traded|listed)",
]
REFERENCE_PATTERNS = [
    r"(?i)(client reference|customer reference|reference (name|contact|detail))",
    r"(?i)(reference (1|2|3|one|two|three))",
    r"(?i)(provide.*(reference|testimonial))",
    r"(?i)(case stud(y|ies))",
    r"(?i)(similar (project|engagement|implementation|client))",
]


# ---------------------------------------------------------------------------
# Sheet Structure Detector
# ---------------------------------------------------------------------------

class SheetStructureDetector:
    """Detects the column layout of an RFP sheet by examining header rows."""

    @staticmethod
    def _match_keyword(cell_value: str, keywords: list[str]) -> bool:
        """Check if cell text matches any keyword (case-insensitive)."""
        if not cell_value:
            return False
        val = cell_value.strip().lower()
        for kw in keywords:
            if kw in val:
                return True
        return False

    @classmethod
    def detect(cls, ws, sheet_name: str = "") -> Optional[dict]:
        """Detect the column layout of a worksheet.

        Returns:
            {
                "id_col": str or None (column letter),
                "question_col": str (column letter),
                "response_col": str (column letter),
                "score_col": str or None (column letter),
                "additional_info_col": str or None (column letter),
                "header_row": int (1-indexed),
                "first_data_row": int (1-indexed),
            }
            or None if no answerable structure detected.
        """
        result = cls._detect_by_headers(ws)
        if result and result.get("question_col") and result.get("response_col"):
            return result

        # Fallback: infer from sheet name patterns
        return cls._detect_by_sheet_name(ws, sheet_name)

    @classmethod
    def _detect_by_headers(cls, ws) -> Optional[dict]:
        """Scan first 15 rows for header keywords."""
        max_scan_rows = min(15, ws.max_row or 1)
        max_scan_cols = min(15, ws.max_column or 1)

        best_header_row = None
        detected = {}

        for row_idx in range(1, max_scan_rows + 1):
            row_detections = {}
            for col_idx in range(1, max_scan_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = str(cell.value).strip() if cell.value is not None else ""
                if not val:
                    continue

                col_letter = get_column_letter(col_idx)

                if cls._match_keyword(val, RESPONSE_KEYWORDS):
                    row_detections["response_col"] = col_letter
                elif cls._match_keyword(val, QUESTION_KEYWORDS):
                    row_detections["question_col"] = col_letter
                elif cls._match_keyword(val, ID_KEYWORDS):
                    row_detections["id_col"] = col_letter
                elif cls._match_keyword(val, SCORE_KEYWORDS):
                    row_detections["score_col"] = col_letter
                elif cls._match_keyword(val, ["additional", "comments", "notes", "remarks", "elaboration"]):
                    row_detections["additional_info_col"] = col_letter

            # If this row has both question and response columns, it's likely the header
            if "question_col" in row_detections and "response_col" in row_detections:
                if best_header_row is None:
                    best_header_row = row_idx
                    detected = row_detections

        if best_header_row is not None:
            detected["header_row"] = best_header_row
            detected["first_data_row"] = best_header_row + 1
            detected.setdefault("id_col", None)
            detected.setdefault("score_col", None)
            detected.setdefault("additional_info_col", None)
            return detected

        return None

    @classmethod
    def _detect_by_sheet_name(cls, ws, sheet_name: str) -> Optional[dict]:
        """Fallback: infer column layout from sheet name patterns."""
        name_lower = sheet_name.lower().strip()

        # Check for "Instructions" or similar non-answerable sheets
        if any(kw in name_lower for kw in ["instruction", "guide", "overview", "summary", "scoring"]):
            return None

        # Detect patterns from Citco-style RFPs
        # Sheet D pattern: ID=A, Question=B, Score=C, Response=D
        if name_lower.startswith("d") and "functional" in name_lower:
            return {
                "id_col": "A", "question_col": "B", "score_col": "C",
                "response_col": "D", "additional_info_col": "E",
                "header_row": cls._find_data_start(ws, "A"),
                "first_data_row": cls._find_data_start(ws, "A") + 1,
            }

        # Sheet E pattern: ID=B, Question=C, Score=D, Response=E
        if name_lower.startswith("e") and "non-functional" in name_lower:
            return {
                "id_col": "B", "question_col": "C", "score_col": "D",
                "response_col": "E", "additional_info_col": "F",
                "header_row": cls._find_data_start(ws, "B"),
                "first_data_row": cls._find_data_start(ws, "B") + 1,
            }

        # Default pattern: ID=B, Question=C, Response=D
        header_row = cls._find_data_start(ws, "B")
        if header_row:
            return {
                "id_col": "B", "question_col": "C", "response_col": "D",
                "score_col": None, "additional_info_col": "E",
                "header_row": header_row,
                "first_data_row": header_row + 1,
            }

        return None

    @staticmethod
    def _find_data_start(ws, id_col_letter: str) -> int:
        """Find the header row by looking for the first row with data in the ID column area."""
        col_idx = column_index_from_string(id_col_letter)
        for row_idx in range(1, min(20, (ws.max_row or 1) + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = str(cell.value).strip() if cell.value is not None else ""
            if val and any(kw in val.lower() for kw in ID_KEYWORDS + QUESTION_KEYWORDS):
                return row_idx
        # Default: assume row 3 is header (common in RFPs with title rows)
        return 3


# ---------------------------------------------------------------------------
# Question Extractor
# ---------------------------------------------------------------------------

class QuestionExtractor:
    """Extracts questions from a worksheet given its detected structure."""

    @staticmethod
    def classify_question(text: str) -> str:
        """Classify question type based on text patterns."""
        if not text:
            return "narrative"

        for pattern in COMPANY_INFO_PATTERNS:
            if re.search(pattern, text):
                return "company_info"

        for pattern in REFERENCE_PATTERNS:
            if re.search(pattern, text):
                return "reference"

        for pattern in BINARY_PATTERNS:
            if re.search(pattern, text):
                return "binary"

        for pattern in NARRATIVE_PATTERNS:
            if re.search(pattern, text):
                return "narrative"

        # Default: if short and statement-like, treat as binary; else narrative
        if len(text) < 100 and not text.endswith("?"):
            return "binary"
        return "narrative"

    @staticmethod
    def is_category_header(row_data: dict, ws, row_idx: int, structure: dict) -> bool:
        """Determine if a row is a category header (not a question)."""
        question_text = row_data.get("question", "")
        id_value = row_data.get("id", "")

        # If there's a proper question ID (e.g., "D.1", "E.15"), it's a question
        if id_value and re.match(r'^[A-Z]\.\d+', str(id_value).strip()):
            return False

        # No ID but has text in question column - likely a category
        if not id_value and question_text:
            # Check if it's short (typical of category headers)
            if len(question_text.strip()) < 80:
                return True
            # Check bold formatting
            q_col_idx = column_index_from_string(structure["question_col"])
            cell = ws.cell(row=row_idx, column=q_col_idx)
            if cell.font and cell.font.bold:
                return True

        # Check for merged cells in this row
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row <= row_idx <= merged_range.max_row:
                # If the merge spans 3+ columns, it's likely a category header
                if merged_range.max_col - merged_range.min_col >= 2:
                    return True

        return False

    @classmethod
    def extract(cls, ws, structure: dict) -> dict:
        """Extract all questions from a worksheet.

        Returns:
            {
                "total_questions": int,
                "categories": [str],
                "questions": [
                    {
                        "row": int,
                        "id": str,
                        "category": str,
                        "question": str,
                        "question_type": str,
                        "current_response": str,
                        "response_col_letter": str,
                        "score_col_letter": str or None,
                    }
                ]
            }
        """
        questions = []
        categories = []
        current_category = "General"

        id_col_idx = column_index_from_string(structure["id_col"]) if structure.get("id_col") else None
        q_col_idx = column_index_from_string(structure["question_col"])
        resp_col_idx = column_index_from_string(structure["response_col"])
        score_col_idx = column_index_from_string(structure["score_col"]) if structure.get("score_col") else None
        addl_col_idx = column_index_from_string(structure["additional_info_col"]) if structure.get("additional_info_col") else None

        first_data_row = structure.get("first_data_row", 2)

        for row_idx in range(first_data_row, (ws.max_row or first_data_row) + 1):
            # Read cell values
            id_val = ""
            if id_col_idx:
                cell = ws.cell(row=row_idx, column=id_col_idx)
                id_val = str(cell.value).strip() if cell.value is not None else ""

            q_cell = ws.cell(row=row_idx, column=q_col_idx)
            question_text = str(q_cell.value).strip() if q_cell.value is not None else ""

            resp_cell = ws.cell(row=row_idx, column=resp_col_idx)
            current_response = str(resp_cell.value).strip() if resp_cell.value is not None else ""

            # Skip empty rows
            if not question_text and not id_val:
                continue

            # Skip rows that look like totals or formulas
            if id_val.lower().startswith("total") or question_text.lower().startswith("total"):
                continue
            # Skip if the question cell contains a formula
            if question_text.startswith("="):
                continue

            row_data = {
                "id": id_val,
                "question": question_text,
                "current_response": current_response,
            }

            # Check if this is a category header
            if cls.is_category_header(row_data, ws, row_idx, structure):
                # Use whichever column has the text
                cat_text = question_text if question_text else id_val
                if cat_text:
                    current_category = cat_text.strip()
                    if current_category not in categories:
                        categories.append(current_category)
                continue

            # It's a question row
            if not question_text:
                continue

            # Read additional info if available
            additional_info = ""
            if addl_col_idx:
                addl_cell = ws.cell(row=row_idx, column=addl_col_idx)
                additional_info = str(addl_cell.value).strip() if addl_cell.value is not None else ""

            question_entry = {
                "row": row_idx,
                "id": id_val,
                "category": current_category,
                "question": question_text,
                "additional_info": additional_info,
                "question_type": cls.classify_question(question_text),
                "current_response": current_response,
                "response_col_letter": structure["response_col"],
            }

            if structure.get("score_col"):
                question_entry["score_col_letter"] = structure["score_col"]
                # Read current score
                if score_col_idx:
                    score_cell = ws.cell(row=row_idx, column=score_col_idx)
                    # Check if it's a formula - don't include formulas
                    if score_cell.data_type == 'f':
                        question_entry["score_is_formula"] = True
                    else:
                        current_score = str(score_cell.value).strip() if score_cell.value is not None else ""
                        question_entry["current_score"] = current_score

            questions.append(question_entry)

        return {
            "total_questions": len(questions),
            "categories": categories,
            "questions": questions,
        }


# ---------------------------------------------------------------------------
# Answer Writer
# ---------------------------------------------------------------------------

class AnswerWriter:
    """Writes answers back into the Excel file, preserving formatting."""

    @staticmethod
    def _get_writable_cell(ws, row: int, col_letter: str):
        """Get a writable cell, handling merged cells by finding the anchor cell.

        If the target cell is a MergedCell, unmerge the range first so we can
        write to it, then write to the anchor cell.
        Returns (cell, was_merged) tuple.
        """
        col_idx = column_index_from_string(col_letter)
        cell = ws.cell(row=row, column=col_idx)

        if not isinstance(cell, MergedCell):
            return cell, False

        # Find the merged range containing this cell
        for merged_range in list(ws.merged_cells.ranges):
            if (merged_range.min_row <= row <= merged_range.max_row and
                    merged_range.min_col <= col_idx <= merged_range.max_col):
                # Unmerge this range so we can write
                ws.unmerge_cells(str(merged_range))
                # Now get the cell again (should be writable)
                cell = ws.cell(row=row, column=col_idx)
                return cell, True

        # Shouldn't reach here, but return the cell anyway
        return ws.cell(row=row, column=col_idx), False

    @staticmethod
    def write(input_path: str, answers_data: dict, output_path: Optional[str] = None) -> str:
        """Write answers back to the Excel file.

        Args:
            input_path: Path to original Excel file
            answers_data: {"file": "...", "answers": {"Sheet Name": [{"row": int, "response_col_letter": str, "answer": str, "score": str?, "score_col_letter": str?}]}}
            output_path: Optional output path (default: input_answered.xlsx)

        Returns:
            Path to the output file
        """
        if output_path is None:
            base, ext = os.path.splitext(input_path)
            output_path = f"{base}_answered{ext}"

        # Open workbook in full read-write mode (preserves formatting)
        wb = load_workbook(input_path)

        total_written = 0
        total_skipped = 0
        total_unmerged = 0

        for sheet_name, sheet_answers in answers_data.get("answers", {}).items():
            if sheet_name not in wb.sheetnames:
                print(f"  WARNING: Sheet '{sheet_name}' not found, skipping", file=sys.stderr)
                continue

            ws = wb[sheet_name]

            for answer in sheet_answers:
                row = answer["row"]
                col_letter = answer["response_col_letter"]

                # Get writable cell (handles merged cells)
                cell, was_merged = AnswerWriter._get_writable_cell(ws, row, col_letter)
                if was_merged:
                    total_unmerged += 1

                cell.value = answer["answer"]
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                total_written += 1

                # Write score if provided
                if answer.get("score") and answer.get("score_col_letter"):
                    score_col = answer["score_col_letter"]
                    score_cell, score_merged = AnswerWriter._get_writable_cell(ws, row, score_col)
                    if score_merged:
                        total_unmerged += 1

                    # Don't overwrite formulas
                    if hasattr(score_cell, 'data_type') and score_cell.data_type == 'f':
                        print(f"  Skipping score for row {row} (formula cell)", file=sys.stderr)
                        total_skipped += 1
                    else:
                        score_cell.value = answer["score"]

        wb.save(output_path)
        wb.close()

        print(f"  Answers written: {total_written}", file=sys.stderr)
        if total_unmerged:
            print(f"  Merged cells unmerged for writing: {total_unmerged}", file=sys.stderr)
        if total_skipped:
            print(f"  Score cells skipped (formulas): {total_skipped}", file=sys.stderr)
        print(f"  Output: {output_path}", file=sys.stderr)

        return output_path


# ---------------------------------------------------------------------------
# Main Orchestrator
# ---------------------------------------------------------------------------

class ExcelRFPParser:
    """Main orchestrator for Excel RFP parsing and answer writing."""

    def list_sheets(self, file_path: str) -> dict:
        """List all sheets with metadata and detected structure."""
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheets = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Basic info
            sheet_info = {
                "name": sheet_name,
                "total_rows": ws.max_row or 0,
                "total_columns": ws.max_column or 0,
            }

            # Try to detect structure
            # Need to reopen without read_only for merged cell detection
            wb_full = load_workbook(file_path, data_only=True)
            ws_full = wb_full[sheet_name]
            structure = SheetStructureDetector.detect(ws_full, sheet_name)
            wb_full.close()

            if structure:
                sheet_info["is_answerable"] = True
                sheet_info["structure"] = {
                    "id_col": structure.get("id_col"),
                    "question_col": structure["question_col"],
                    "response_col": structure["response_col"],
                    "score_col": structure.get("score_col"),
                    "header_row": structure["header_row"],
                }

                # Count questions (quick scan)
                q_col_idx = column_index_from_string(structure["question_col"])
                first_data = structure.get("first_data_row", 2)
                q_count = 0
                for row_idx in range(first_data, (ws.max_row or first_data) + 1):
                    cell = ws.cell(row=row_idx, column=q_col_idx)
                    val = str(cell.value).strip() if cell.value is not None else ""
                    if val and not val.startswith("=") and len(val) > 5:
                        q_count += 1
                sheet_info["question_count"] = q_count
            else:
                sheet_info["is_answerable"] = False
                sheet_info["question_count"] = 0
                sheet_info["note"] = "No answerable question structure detected"

            sheets.append(sheet_info)

        wb.close()

        return {
            "file": os.path.basename(file_path),
            "file_path": file_path,
            "sheets": sheets,
        }

    def extract_questions(self, file_path: str, sheet_names: list[str]) -> dict:
        """Extract questions from specified sheets."""
        wb = load_workbook(file_path, data_only=True)
        result = {
            "file": os.path.basename(file_path),
            "file_path": file_path,
            "extracted_at": datetime.now().isoformat(),
            "sheets": {},
        }

        for sheet_name in sheet_names:
            # Find matching sheet (allow partial matching)
            matched_name = self._match_sheet_name(sheet_name, wb.sheetnames)
            if not matched_name:
                print(f"  WARNING: No sheet matching '{sheet_name}', skipping", file=sys.stderr)
                continue

            ws = wb[matched_name]
            structure = SheetStructureDetector.detect(ws, matched_name)

            if not structure:
                print(f"  WARNING: Could not detect structure for '{matched_name}', skipping", file=sys.stderr)
                continue

            extraction = QuestionExtractor.extract(ws, structure)

            result["sheets"][matched_name] = {
                "structure": {
                    "id_col": structure.get("id_col"),
                    "question_col": structure["question_col"],
                    "response_col": structure["response_col"],
                    "score_col": structure.get("score_col"),
                },
                **extraction,
            }

            print(f"  {matched_name}: {extraction['total_questions']} questions in {len(extraction['categories'])} categories", file=sys.stderr)

        wb.close()
        return result

    def write_answers(self, file_path: str, answers_json_path: str, output_path: Optional[str] = None) -> str:
        """Write answers back to the Excel file."""
        with open(answers_json_path, "r") as f:
            answers_data = json.load(f)

        return AnswerWriter.write(file_path, answers_data, output_path)

    @staticmethod
    def _match_sheet_name(query: str, available: list[str]) -> Optional[str]:
        """Match a sheet name query to available sheet names (exact or partial)."""
        query_clean = query.strip()

        # Exact match
        for name in available:
            if name == query_clean:
                return name

        # Case-insensitive match
        for name in available:
            if name.lower() == query_clean.lower():
                return name

        # Partial match (query is a prefix or substring)
        for name in available:
            if query_clean.lower() in name.lower():
                return name

        # Single letter match (e.g., "D" matches "D. Functional Requirements")
        if len(query_clean) <= 2:
            for name in available:
                if name.lower().startswith(query_clean.lower()):
                    return name

        return None


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="RFP/RFI Excel Questionnaire Parser and Answer Writer"
    )
    parser.add_argument(
        "--input", required=True,
        help="Path to the Excel (.xlsx) file"
    )
    parser.add_argument(
        "--parse-only", action="store_true",
        help="Parse mode: extract questions as JSON to stdout"
    )
    parser.add_argument(
        "--sheets", type=str, default=None,
        help="Comma-separated sheet names to extract (for --parse-only). If omitted, lists all sheets."
    )
    parser.add_argument(
        "--write-answers", type=str, default=None,
        help="Write mode: path to answers JSON file to write back into the Excel"
    )
    parser.add_argument(
        "--output-file", type=str, default=None,
        help="For --write-answers: output Excel file path (default: <input>_answered.xlsx)"
    )
    parser.add_argument(
        "--verbose", action="store_true",
        help="Enable verbose logging"
    )

    args = parser.parse_args()

    # Setup logging
    level = logging.DEBUG if args.verbose else logging.WARNING
    logging.basicConfig(level=level, format="%(levelname)s: %(message)s", stream=sys.stderr)

    # Validate input file
    if not os.path.isfile(args.input):
        print(f"ERROR: File not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    rfp_parser = ExcelRFPParser()

    if args.parse_only:
        if args.sheets:
            # Extract questions from specified sheets
            sheet_names = [s.strip() for s in args.sheets.split(",")]
            print(f"Extracting questions from: {', '.join(sheet_names)}", file=sys.stderr)
            result = rfp_parser.extract_questions(args.input, sheet_names)
        else:
            # List all sheets
            print(f"Listing sheets in: {os.path.basename(args.input)}", file=sys.stderr)
            result = rfp_parser.list_sheets(args.input)

        # Output JSON to stdout
        json.dump(result, sys.stdout, indent=2, ensure_ascii=False)
        print()  # trailing newline

    elif args.write_answers:
        if not os.path.isfile(args.write_answers):
            print(f"ERROR: Answers file not found: {args.write_answers}", file=sys.stderr)
            sys.exit(1)

        print(f"Writing answers from: {args.write_answers}", file=sys.stderr)
        output = rfp_parser.write_answers(args.input, args.write_answers, args.output_file)
        print(json.dumps({"status": "success", "output_file": output}))

    else:
        parser.print_help()
        sys.exit(1)


if __name__ == "__main__":
    main()
