import io
import logging

from openpyxl import load_workbook

from app.documents.parsers.base import BaseParser, ParsedDocument

logger = logging.getLogger(__name__)


class XlsxParser(BaseParser):
    """Microsoft Excel (.xlsx) parser."""

    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xls"]

    def parse(self, file_data: bytes, filename: str) -> ParsedDocument:
        try:
            wb = load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
        except Exception as e:
            logger.error(f"Failed to parse XLSX {filename}: {e}")
            raise

        all_text_parts = []
        all_tables = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_text_parts.append(f"--- Sheet: {sheet_name} ---")

            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(cell) if cell is not None else "" for cell in row]
                if any(c.strip() for c in cells):  # skip completely empty rows
                    rows.append(cells)
                    all_text_parts.append(" | ".join(cells))

            if rows:
                all_tables.append(rows)

        wb.close()

        full_text = "\n".join(all_text_parts)
        return ParsedDocument(
            text=full_text,
            page_count=len(wb.sheetnames),
            tables=all_tables,
            metadata={
                "filename": filename,
                "parser": "xlsx",
                "sheet_names": wb.sheetnames,
            },
        )
